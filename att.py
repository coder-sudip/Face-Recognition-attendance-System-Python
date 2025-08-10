import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

def write_attendance(name, user_id, filename="attendance.xlsx"):
    # Check if file exists
    if not os.path.exists(filename):
        # Create a new workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        # Write headers
        ws["A1"] = "Name"
        ws["B1"] = "ID"
        ws["C1"] = "Timestamp"
    else:
        # Load the existing workbook
        wb = load_workbook(filename)
        ws = wb.active

    # Get current date and time
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Find next empty row
    next_row = ws.max_row + 1

    # Write the data
    ws[f"A{next_row}"] = name
    ws[f"B{next_row}"] = user_id
    ws[f"C{next_row}"] = timestamp

    # Save the workbook
    wb.save(filename)
    # print(f"Attendance saved for {name} ({user_id}) at {timestamp}")

# Example usage
if __name__ == "__main__":
    write_attendance("Sudip", "105")
